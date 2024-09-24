"""
Authored by James Gaskell

08/19/2024

Edited by:

"""

import pandas as pd
import openpyxl, easygui
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def file_open_check(filepath):
    wb = openpyxl.load_workbook(filepath)
    try:
        wb.save(filepath)
        wb.close()
        return False
    except:
        return True


"""Uses pandas to read excel document selected by the user
Args:
    String: Filepath decide dby the user in the selection process
Returns:
    Dict: Dataframes for each sheet in the file
    List[String]:Sheetnames in the file
"""
def get_dataFrames(filepath):
    xl_file = pd.ExcelFile(filepath)
    dfs = pd.read_excel(xl_file, sheet_name=None)
    sheets = xl_file.sheet_names

    xl_file.close()
    return dfs,sheets

def write_dataframes(dfs, sheetnames, filepath):
    xl_writer = pd.ExcelWriter("filepath", engine='openpyxl')
    for sheet in sheetnames:
         dfs[sheet].to_excel(xl_writer, sheet_name=sheet, index=False)
    xl_writer.save()
    return True


def set_field_format(ws, column_name, column_index):
    if column_name == "date_created":
        for cell in ws[get_column_letter(column_index+1)]:
             cell.alignment = Alignment(horizontal='right')
             cell.number_format = "YYYY-MM-DD"
    return True


def df_to_excel(dfs, sheetnames, filepath):

    wb = openpyxl.load_workbook(filepath)  # load as openpyxl workbook; useful to keep the original layout
                                 # which is discarded in the following dataframe
    
    for sheet in sheetnames:
         
        ws = wb[sheet]

        for index, column_name in enumerate(list(dfs[sheet].columns.values)): #Writes the column headings to the file
             ws.cell(1, index+1).value = column_name

             
             if column_name == ("date_created"):
                  set_field_format(ws, "date_created", index)


        for r_idx, row in dfs[sheet].iterrows():
             for c_idx, value in enumerate(row, 1):
                  ws.cell(row=r_idx+2, column=c_idx).value = value

    wb.save(filepath)
    wb.close()
