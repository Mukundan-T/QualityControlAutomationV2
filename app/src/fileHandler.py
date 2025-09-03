"""
Authored by James Gaskell

01/24/2025

Edited by: Mukundan Thanigaivelan

"""
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import messagebox

ACCEPTED_EXTENSIONS = ('csv', 'xlsx', 'xls')
DEFAULT_EXT = 'xlsx'


"""Openpyxl does not have permission to edit an open file
    This function checks if the file is open in an editor so an exception can be raised
Args:
    filepath: the path to the excel file
Returns:
    Boolean: True if the file is open in an editor
"""
def file_open_check(filepath):
    wb = openpyxl.load_workbook(filepath)
    try:
        wb.save(filepath)
        wb.close()
        return False
    except:
        return True
    

#Computationally expensive but should work for now
"""Removes specific highlight colors from the spreadsheet to allow the program to be run
    continually after each error is rectified
Args:
    ExcelFile: Excel file to remove colors on - contains the error colors dictionary
    wb: Work book opened with Openpyxl
    colors_to_remove: allows specification of individual colors so running spreadsheetChecks alone
        (for example) doesn't remove failure highlighting
"""
def reset_colors(ExcelFile, wb, colors_to_remove):
    fill_reset = openpyxl.styles.PatternFill(fill_type=None)
    for sheet in ExcelFile.sheetList:
        ws = wb[sheet.sheetName]
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill.start_color.index in colors_to_remove.values():
                    cell.fill = fill_reset


"""Highlights rows with errors and/or failures with their corresponding hex color
    Since the function pulls from both the error dictionary and the failure dictionary
    it can be used in both parts of the program
Args:
    ExcelFile: Contains the failure/error inormation and the highliting colors
Returns:
    Boolean: True if the save is successful, False if not
"""
def highlight_errors(ExcelFile):

    xl_file = pd.ExcelFile(ExcelFile.filePath, engine="openpyxl")       
    wb = openpyxl.load_workbook(xl_file)

    ExcelFile.retrieveColorCache()

    reset_colors(ExcelFile, wb, ({**ExcelFile.cachedColors, **ExcelFile.errorColors}))

    for sheet in ExcelFile.sheetList:
        dt = pd.read_excel(xl_file, sheet.sheetName)
        ws = wb[sheet.sheetName]

        errors = sheet.getSheetErrorDict()
        errors.update(sheet.getSheetFailureDict())

        colors = ExcelFile.errorColors
        colors.update (ExcelFile.failColors)

        for file in errors:
            error_color = colors[errors[file]]
        
            if error_color != None:
                 fill = openpyxl.styles.PatternFill(start_color=error_color, end_color=error_color, fill_type="solid")
                 for index, row in dt.iterrows():
                    try:
                        if file == dt['Filename'][index]:
                            for y in range(1, ws.max_column+1):
                                ws.cell(row=index+2, column=y).fill = fill
                    except:
                        pass

    try: 
        wb.save(ExcelFile.filePath)
        wb.close()
        xl_file.close()
        return True
    except:
        wb.close()
        xl_file.close()
        return False
    

"""Currently used to set date format to the correct ISO
    Can be expanded to include formatting options for other fields
Args:
    ws: the worksheet object
    column_name: the name of the column
    column_index: the index of the column
"""
def set_field_format(ws, column_name, column_index):
    if column_name == "date_created":
        for cell in ws[get_column_letter(column_index+1)]:
             cell.alignment = Alignment(horizontal='right')
             cell.number_format = "YYYY-MM-DD"


"""Writes the edited excelfile over the original version
    Resets the original colors of the spreadsheet based on current error/fail colors
    and the cached colors saved to file. Sets the date format to the requires ISO 
    for upload to ARCHES
Args:
    ExcelFile: the current file after all QC changes were made
Returns:
    True: if the save is completed
    False: if the file is open in editor so the user can be made aware with an error message
"""
def write_excelfile(ExcelFile):
    wb = openpyxl.load_workbook(ExcelFile.filePath)

    ExcelFile.retrieveColorCache()
 
    reset_colors(ExcelFile, wb, ({**ExcelFile.cachedColors, **ExcelFile.failColors})) # Only removes fail colors since this is independent of spreadsheetChecks

    for sheet in ExcelFile.sheetList:
        ws = wb[sheet.sheetName]

        for index, column_name in enumerate(list(ExcelFile.dataFrames[sheet.sheetName].columns.values)):
            ws.cell(1, index+1).value = column_name

            # This is where we call the field formatter
            # Any expansion to formatter requires an additional call here
            if column_name == ("date_created"):
                set_field_format(ws, "date_created", index)

        for r_idx, row in ExcelFile.dataFrames[sheet.sheetName].iterrows():
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx+2, column=c_idx).value = value

    try:
        wb.save(ExcelFile.filePath)
        wb.close()
        return True
    except:
        return False
    

"""Determines whether the filename selected by the user already has a file extension
Args:
    filename: the name selected by the user in the GUI
Returns:
    True if this extension is an excel file
    False if the file does not contain an extension
    ValueError if the file has an extension but it is not in the proper format
"""
def extract_ext(filename):
    if '.' in filename:
        ext = filename.split('.')[-1]
        if ext in ACCEPTED_EXTENSIONS:
            return True
        else:
            raise ValueError(f'File extension must be one of: {ACCEPTED_EXTENSIONS}')
    else:
        return False

"""Simple Tkinter window opened if the file already exists
    Asks the user if they would like to overwrite
    Returns the choice which causes the spreadsheet creation to proceed or cancel
"""
def ask_yes_cancel(title="Confirm", message="This file already exists. Would you like to overwrite?"):
    """Simple """
    root = tk.Tk()
    root.withdraw()
    result = messagebox.askyesnocancel(title, message)
    root.destroy()
    return result

"""Creates a spreadsheet based on the template at Union College
    Adds as many sheets as requested by the user with their respective names
    Adds the fields currently used for digitization as headers
    Resizes the columns to improve readability
"""
def generateSpreadsheet(filename, sheetnames):
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    try:
        if extract_ext(filename):
            filepath = os.path.join(desktop_path, filename)
        else:
            filepath = os.path.join(desktop_path, (filename + ".xlsx"))
    except ValueError as e:
        messagebox.showerror("Error", str(e))
        return False

    if os.path.exists(filepath):
        result = ask_yes_cancel()
        if not result:
            return False

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    headers = ['ismemberof', 'type', 'aspace_id', 'documents', 'local_identifier', 'aspaceTitle',
               'label (title)', 'description', 'titleProper', 'creator', 'creator_role', 'contributor',
               'contributor_role', 'subjects_personal_names','Language', 'publisher',
               'date_created_free', 'date_created', 'genre', 'rights_statements', 
               'extent (total page count including covers)', 'physical_location', 
               'Scanning Instructions', 'Filename', 'date_digital','Scanner Initials',
               'QC Pass/Fail', 'QC Initials', 'QC Comments']
    
    column_widths = {
        'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 28, 'H': 40,'I': 12,
        'J': 70, 'K': 30, 'L': 12, 'M': 22, 'N': 15, 'O': 27, 'P': 32, 'Q': 32, 'R': 35, 
        'S': 37, 'T': 14, 'U': 40, 'V': 23, 'W': 25, 'X': 23, 'Y': 17, 'Z': 23,
        'AA': 23, 'AB': 23, 'AC': 23,
    }

    for name in sheetnames:
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

    wb.save(filepath)
    
    return True, filepath